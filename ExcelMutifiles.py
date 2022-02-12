import openpyxl
import pathlib
from FolderCreator import Folder

class ExcelCopy:
    """class that creates a copy of existing format file wile keeping all hidden sheets.
    Can create more than one sheet in file with a same format"""

    def __init__(self, source_file, **kwargs):
        self.source_file = source_file
        self.source_worksheet_name = kwargs.get("worksheet", None)
        self.target_folder = kwargs.get("folder", "./")
        self.target_file_name = kwargs.get("name", f"{self.target_folder}/{self.target_folder}.xlsx")

    def __enter__(self):
        self.source_wb = openpyxl.load_workbook(self.source_file)
        if self.source_worksheet_name:
            self.source_ws = self.source_wb[self.source_worksheet_name]
        else:
            self.source_ws = self.source_wb.active

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.source_wb.save(self.target_folder)

    def create_sheet_copy(self, greed_lines=False, **kwargs):
        """creates a copy and delete the sample sheet"""
        new_worksheet = self.source_wb.copy_worksheet(self.source_ws)
        new_worksheet.title = kwargs.get("title", None)
        new_worksheet.sheet_properties.tabColor = kwargs.get("title_color", None)
        new_worksheet.sheet_view.showGridLines = greed_lines

    def create_workbook(self, workbook_name, *args):
        """create file according to format, for multiple sheet provide sheet names as list to *args"""

    def check_type_source_file(self):
        """checks that source file is excel"""

    # @Folder.create_folder
    def check_for_folder(self):
        """checks if folder does exist. use a decorator to other class if not and create"""

file = r"C:\Users\marin\PycharmProjects\Naftya OOP\Format.xlsx"
with ExcelCopy(file) as ec:
    ec.create_sheet_copy()
